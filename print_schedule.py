""" Functions used to output various parts of the schedule
"""
from io import TextIOWrapper
import typing
import logging
from openpyxl import Workbook
from class_defs import League, Schedule


def print_schedule(
    sch: Schedule,
    leag_list: list[str],
    outfile: typing.Union[str, TextIOWrapper] = "print",
):
    """Print compact schedule"""
    if outfile == "print":

        def pline(txt):
            print(txt)

    else:

        def pline(txt):
            outfile.write(txt)

    for rnd in range(sch.nrounds):
        _rnd = sch.rounds[rnd]
        line = f"{_rnd.play_date}"
        for match in range(_rnd.nmatches):
            _match = _rnd.matches[match]
            _away = _match.away
            _home = _match.home
            line = (
                line
                + " "
                + f" {leag_list.index(_away) + 1:2} @ {leag_list.index(_home) + 1:2}"
            )
        pline(line)


def print_league(
    leag: League,
    leag_list: list[str],
    outfile: typing.Union[str, TextIOWrapper] = "print",
):
    """Print list of teams"""
    if outfile == "print":

        def pline(txt):
            print(txt)

    else:

        def pline(txt):
            outfile.write(txt + "\n")

    line = ""
    max_teams = 0
    logging.info("Total number of divisions: %d", leag.ndivs)
    for _div_i in range(leag.ndivs):
        logging.info("Division %d in loop out of %s", _div_i, leag.ndivs)
        line = line + f"{leag.divs[_div_i].name:<43}"
        max_teams = max(max_teams, leag.divs[_div_i].nteams)
    pline(line)
    line = ""
    for _team_i in range(max_teams):
        for _div_i in range(leag.ndivs):
            _div = leag.divs[_div_i]
            if _team_i < _div.nteams:
                _num = str(leag_list.index(_div.teams[_team_i]) + 1)
                _full_name = _div.teams[_team_i] + " - " + _div.capts[_team_i]
                line = line + f"{_num:>2}. {_full_name:<39}"
            else:
                line = line + f"{'':<32}"
        pline(line)
        line = ""


def print_team_schedule(
    sch: Schedule,
    team: str,
    team_list: list[str],
    capt_list: list[str],
    outfile: typing.Union[str, TextIOWrapper] = "print",
):

    """Print individual team schedule"""
    if outfile == "print":

        def pline(txt):
            print(txt)

    else:

        def pline(txt):
            outfile.write(txt + "\n")

    line = ""

    pline("\nTeam: " + team + "\n")
    for rnd in range(sch.nrounds):
        _rnd = sch.rounds[rnd]
        line = f"{_rnd.play_date}"
        game_not_found = True
        match = 0
        while game_not_found and match < _rnd.nmatches:
            _match = _rnd.matches[match]
            if _match.home == team:
                _teamidx = team_list.index(_match.away)
                _capt = capt_list[_teamidx]
                if "Bye" not in _match.away:
                    line = line + f"  vs. {_match.away} ({_capt})"
                else:
                    line = line + "  --- BYE ---"
                game_not_found = False
            elif _match.away == team:
                _teamidx = team_list.index(_match.home)
                _capt = capt_list[_teamidx]
                if "Bye" not in _match.home:
                    line = line + f"  @ {_match.home} ({_capt})"
                else:
                    line = line + "  --- BYE ---"
                game_not_found = False
            else:
                match = match + 1
        if game_not_found:
            logging.warning("Bye week is not expected.")
            line = line + "Bye Week"
        pline(line)


def create_excel_workbook(
    file_name: str,
    league: League,
    schedule: Schedule,
    list_teams: list[str],
    list_capts: list[str],
):
    """Create standard IHPL schedule worksheet"""
    # write to excel
    wb = Workbook()

    # write to active worksheet
    ws = wb.active

    # Title
    ws["A1"] = league.name
    # First half of schedule
    for rnd_num in range(int(schedule.nrounds / 2)):
        sch_round = schedule.rounds[rnd_num]
        row = [
            sch_round.play_date,
        ]
        for match in range(sch_round.nmatches):
            _match = sch_round.matches[match]
            _away = _match.away
            _home = _match.home
            row.append(
                f"{str(list_teams.index(_away) + 1).zfill(2):2} @ {str(list_teams.index(_home) + 1).zfill(2):2}"
            )
        ws.append(row)
    ws["B17"] = league.divs[0].name
    ws["G17"] = league.divs[1].name
    for _team_i in range(league.divs[0].nteams):
        row = [""]
        row.append(_team_i + 1)
        row.append(f"{list_teams[_team_i]} - {list_capts[_team_i]}")
        row.extend(["", "", ""])
        row.append(_team_i + 11)
        row.append(f"{list_teams[_team_i + 10]} - {list_capts[_team_i + 10]}")
        ws.append(row)

    # Title
    ws["A29"] = league.name
    # Second half of schedule
    for rnd_num in range(int(schedule.nrounds / 2), schedule.nrounds):
        sch_round = schedule.rounds[rnd_num]
        row = [
            sch_round.play_date,
        ]
        for match in range(sch_round.nmatches):
            _match = sch_round.matches[match]
            _away = _match.away
            _home = _match.home
            row.append(
                f"{str(list_teams.index(_away) + 1).zfill(2):2} @ {str(list_teams.index(_home) + 1).zfill(2):2}"
            )
        ws.append(row)
    ws["B45"] = league.divs[0].name
    ws["G45"] = league.divs[1].name
    for _team_i in range(league.divs[0].nteams):
        row = [""]
        row.append(_team_i + 1)
        row.append(f"{list_teams[_team_i]} - {list_capts[_team_i]}")
        row.extend(["", "", ""])
        row.append(_team_i + 11)
        row.append(f"{list_teams[_team_i + 10]} - {list_capts[_team_i + 10]}")
        ws.append(row)

    wb.save(file_name)
